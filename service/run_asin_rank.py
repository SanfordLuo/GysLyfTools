"""
@Author  :   luoyafei
@Time    :   2025/7/26 21:53
@Desc    :   跑商品排名
"""
import os
import time
import datetime
import random
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urljoin, urlencode, urlparse
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
import selenium.webdriver.support.expected_conditions as ec
from setting.config import DATA_PATH
from utils.logger import config_log


class RunAsinRank(object):
    def __init__(self):
        self.driver = None
        self.wait_time = 5

    def init_driver(self):
        """
        :return:
        """
        logging.info("init_driver begin")

        try:
            chrome_options = Options()

            # 基础配置, --disable-dev-shm-usage 和 --no-sandbox 参数，这是无头模式在Linux服务器稳定运行的关键配置
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")  # 关键解决内存问题
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--incognito")

            # 高级防检测
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)

            # 自定义UA
            chrome_version = random.randint(120, 137)
            user_agent = f"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_version}.0.0.0 Safari/537.36"
            chrome_options.add_argument(f'user-agent={user_agent}')

            # 初始化驱动
            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

            # 超时设置
            self.driver.set_page_load_timeout(30)
            self.driver.set_script_timeout(20)

            logging.info("init_driver success")
            return True

        except Exception as error:
            logging.exception(f"init_driver error: {error}")

        return False

    def quit_driver(self):
        """
        :return:
        """
        self.driver.quit()

    def open_home_and_change_zipcode(self, zipcode):
        """
        打开首页
        :return:
        """

        for _idx in range(3):
            logging.info(f"open_home begin, {_idx}")

            try:
                # url = "https://www.amazon.com/"
                url = "https://www.amazon.com/errors/validateCaptcha"
                self.driver.get(url)
                time.sleep(self.wait_time)

                # 检测邮编是否更新成功
                try:
                    element = self.driver.find_element(By.XPATH, '//*[@id="glow-ingress-line2"]')
                    element_text = element.text
                    if zipcode in element_text:
                        logging.info("change_zipcode success")
                        return True
                except Exception as error:
                    logging.error("change_zipcode error")

                # 如果有validateCaptcha
                if "validateCaptcha" in self.driver.current_url:
                    try:
                        if len(self.driver.current_url.split("?")) == 1:
                            form_action = "/errors/validateCaptcha"
                            element = self.driver.find_element(By.XPATH, f'//*[@action="{form_action}"]')
                            inputs = element.find_elements(By.TAG_NAME, 'input')
                            params = {}
                            for input_elem in inputs:
                                name = input_elem.get_attribute('name')
                                value = input_elem.get_attribute('value')
                                if name and value:
                                    params[name] = value
                            base_url = self.driver.current_url
                            full_url = urljoin(base_url, form_action)
                            parsed_url = urlparse(full_url)
                            captcha_link = parsed_url._replace(query=urlencode(params)).geturl()
                            logging.info(f"captcha_link: {captcha_link}")
                            self.driver.get(captcha_link)
                            time.sleep(self.wait_time)
                        if "validateCaptcha" in self.driver.current_url:
                            button = self.driver.find_element(By.CLASS_NAME, "a-button-text")
                            button.click()
                            time.sleep(self.wait_time)
                    except Exception as error:
                        logging.error(f"validateCaptcha error: {error}")

                # 有时候进了首页不显示改邮编,需要点一下这个
                try:
                    button = WebDriverWait(self.driver, self.wait_time).until(
                        ec.visibility_of_element_located((By.XPATH, '//*[@href="/ref=nav_bb_logo"]')))
                    button.click()
                    time.sleep(self.wait_time)
                except Exception as error:
                    logging.info(f"no nav_bb_logo")

                # 检测首页进入后是否有修改地址的地方
                WebDriverWait(self.driver, self.wait_time).until(
                    ec.visibility_of_element_located((By.XPATH, '//*[@id="nav-global-location-popover-link"]')))

                logging.info("open_home success")

                # 更改邮编
                is_success = self.change_zipcode(zipcode)
                if is_success:
                    return True

            except Exception as error:
                logging.exception(f"open_home error: {error}")

        return False

    def change_zipcode(self, zipcode):
        """
        更改邮编
        :return:
        """
        logging.info("change_zipcode begin")

        try:
            try:
                button = WebDriverWait(self.driver, self.wait_time).until(
                    ec.visibility_of_element_located((By.XPATH, '//input[@data-action-type="SELECT_LOCATION"]')))
                button.click()
                time.sleep(self.wait_time)
            except Exception as error:
                button = WebDriverWait(self.driver, self.wait_time).until(
                    ec.visibility_of_element_located((By.XPATH, '//*[@id="glow-ingress-line2"]')))
                button.click()
                time.sleep(self.wait_time)

            button = WebDriverWait(self.driver, self.wait_time).until(
                ec.visibility_of_element_located((By.XPATH, '//input[@id="GLUXZipUpdateInput"]')))
            button.click()
            button.send_keys(zipcode)
            time.sleep(self.wait_time)

            button = WebDriverWait(self.driver, self.wait_time).until(
                ec.visibility_of_element_located((By.XPATH, '//*[@id="GLUXZipUpdate"]/span/input')))
            button.click()
            time.sleep(self.wait_time)

            button = WebDriverWait(self.driver, self.wait_time).until(
                ec.presence_of_element_located((By.XPATH, '//*[@id="GLUXConfirmClose"]')))
            self.driver.execute_script("arguments[0].click();", button)
            time.sleep(self.wait_time)

            try:
                button = WebDriverWait(self.driver, self.wait_time).until(
                    ec.presence_of_element_located((By.XPATH, '//*[@name="glowDoneButton"]')))
                button.click()
                time.sleep(self.wait_time)
            except Exception as error:
                logging.error("glowDoneButton error.")

            # 检测邮编是否更新成功
            element = self.driver.find_element(By.XPATH, '//*[@id="glow-ingress-line2"]')
            element_text = element.text
            if zipcode in element_text:
                logging.info("change_zipcode success")
                return True

        except Exception as error:
            logging.exception(f"change_zipcode error: {error}")

        return False

    def parse_asin(self):
        """
        处理商品
        :return:
        """
        try:
            sponsored_list = []
            organic_list = []

            WebDriverWait(self.driver, self.wait_time).until(
                ec.presence_of_element_located((By.XPATH, '//div[@role="listitem" and @data-index]')))
            divs = self.driver.find_elements(By.XPATH, '//div[@role="listitem" and @data-index]')
            for _div in divs:
                # 判断是否是 Sponsored 的
                sponsored_span = None
                try:
                    sponsored_span = _div.find_element(
                        By.XPATH, './/span[@aria-label="View Sponsored information or leave ad feedback"]')
                except Exception as error:
                    logging.info(f"no sponsored_span")
                asin = _div.get_attribute("data-asin")
                if sponsored_span:
                    sponsored_list.append(asin)
                else:
                    organic_list.append(asin)

            sponsored_dict = {str(_idx): _v for _idx, _v in enumerate(sponsored_list)}
            organic_dict = {str(_idx): _v for _idx, _v in enumerate(organic_list)}

            logging.info("parse_asin success")
            logging.info(f"sponsored_dict: {sponsored_dict}")
            logging.info(f"organic_dict: {organic_dict}")
            return sponsored_dict, organic_dict

        except Exception as error:
            logging.exception(f"parse_asin error: {error}")

        return {}, {}

    def search_keywords(self, keywords):
        """
        搜索关键词
        :return:
        """
        logging.info("search_keywords begin")

        try:
            button = WebDriverWait(self.driver, self.wait_time).until(
                ec.visibility_of_element_located((By.XPATH, '//*[@id="twotabsearchtextbox"]')))
            button.click()
            button.send_keys(keywords)
            time.sleep(self.wait_time)

            button = WebDriverWait(self.driver, self.wait_time).until(
                ec.visibility_of_element_located((By.XPATH, '//*[@id="nav-search-submit-button"]')))
            button.click()
            time.sleep(self.wait_time)

            sponsored_dict, organic_dict = self.parse_asin()
            if sponsored_dict or organic_dict:
                logging.info("search_keywords success")
                return sponsored_dict, organic_dict

        except Exception as error:
            logging.exception(f"search_keywords error: {error}")

        return {}, {}

    def search_keywords_next(self, page_num):
        """
        下一页
        :return:
        """
        logging.info(f"search_keywords_next begin. page_num: {page_num}")

        try:
            button = WebDriverWait(self.driver, self.wait_time).until(
                ec.visibility_of_element_located((By.XPATH, f'//*[@aria-label="Go to page {page_num}"]')))
            button.click()
            time.sleep(self.wait_time)

            sponsored_dict, organic_dict = self.parse_asin()
            if sponsored_dict or organic_dict:
                logging.info("search_keywords_next success")
                return sponsored_dict, organic_dict

        except Exception as error:
            logging.exception(f"search_keywords_next error: {error}")

        return {}, {}

    def run_asin(self, zipcode, keywords, total_page_nums):
        all_asin_dict = {}
        is_success = self.init_driver()
        if not is_success:
            return all_asin_dict

        # 打开首页并且更改邮编
        is_success = self.open_home_and_change_zipcode(zipcode)
        if not is_success:
            return all_asin_dict

        # 搜索关键词
        all_asin_dict = {}
        sponsored_dict, organic_dict = self.search_keywords(keywords)
        if not sponsored_dict:
            return all_asin_dict
        all_asin_dict["1"] = {
            "sponsored_dict": sponsored_dict,
            "organic_dict": organic_dict
        }

        # 后面的太多, 只要广告位
        for idx in range(1, total_page_nums):
            page_num = idx + 1
            sponsored_dict, organic_dict = self.search_keywords_next(page_num)
            if not sponsored_dict:
                return all_asin_dict
            all_asin_dict[str(page_num)] = {
                "sponsored_dict": sponsored_dict,
                "organic_dict": organic_dict
            }

        logging.info(f"all_asin_dict: {all_asin_dict}")
        return all_asin_dict

    def save_excel(self, all_asin_dict, xlsx_file):
        """
        保存
        :return:
        """
        data = {
            "rank": [],
            "asin": []
        }
        save_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        data["rank"].append(f"   ")
        data["asin"].append(f"   ")
        data["rank"].append(f"执行时间")
        data["asin"].append(save_time)

        for _page, _page_data in all_asin_dict.items():
            data["rank"].append(f"   ")
            data["asin"].append(f"   ")
            data["rank"].append(f"第{_page}页")
            data["asin"].append(f"广告位")
            sponsored_dict = _page_data.get("sponsored_dict", {})
            for _rank, _asin in sponsored_dict.items():
                data["rank"].append(_rank)
                data["asin"].append(_asin)

            data["rank"].append(f"   ")
            data["asin"].append(f"   ")
            data["rank"].append(f"第{_page}页")
            data["asin"].append(f"自然位")
            organic_dict = _page_data.get("organic_dict", {})
            for _rank, _asin in organic_dict.items():
                data["rank"].append(_rank)
                data["asin"].append(_asin)

        sheet_name = str(datetime.datetime.now().hour)

        # 将数据转换为DataFrame
        df = pd.DataFrame(data)

        # 检查文件是否存在
        if not os.path.exists(xlsx_file):
            # 文件不存在 - 创建新文件并写入数据
            df.to_excel(xlsx_file, sheet_name=sheet_name, index=False)
            logging.info(f"文件 '{xlsx_file}' 已创建，数据写入工作表 '{sheet_name}'")
        else:
            # 文件存在 - 处理工作表
            try:
                # 加载工作簿
                book = load_workbook(xlsx_file)

                # 检查工作表是否存在
                if sheet_name in book.sheetnames:
                    logging.info(f"工作表 '{sheet_name}' 已存在，将在现有数据右侧空一列后添加新数据")

                    # 获取目标工作表
                    ws = book[sheet_name]

                    # 确定新数据的起始列（现有最大列 + 2，中间空一列）
                    if ws.max_column > 0:
                        start_col = ws.max_column + 2  # 空一列
                    else:
                        start_col = 1  # 如果工作表为空，从第一列开始

                    # 写入列标题（第一行）
                    for c_idx, col_name in enumerate(df.columns, start_col):
                        ws.cell(row=1, column=c_idx, value=col_name)

                    # 写入数据（从第二行开始）
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                        for c_idx, value in enumerate(row, start_col):
                            ws.cell(row=r_idx, column=c_idx, value=value)

                    # 保存工作簿
                    book.save(xlsx_file)
                    logging.info(f"成功在 '{sheet_name}' 工作表右侧空一列后添加 {len(df)} 列数据")

                else:
                    # 工作表不存在 - 创建新工作表并写入数据
                    # 使用pandas写入新工作表
                    with pd.ExcelWriter(xlsx_file, mode='a', engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logging.info(f"工作表 '{sheet_name}' 不存在，已创建并写入数据")

            except Exception as e:
                logging.info(f"操作失败: {str(e)}")
                # 回退方案
                try:
                    # 尝试使用pandas直接追加新工作表
                    with pd.ExcelWriter(xlsx_file, mode='a', engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logging.info("使用回退方案成功写入数据")
                except Exception as fallback_error:
                    logging.info(f"回退方案也失败: {str(fallback_error)}")

    def main_run(self, zipcode, keywords, total_page_nums, xlsx_file):
        """
        入口
        :return:
        """
        status = "failed"
        for _idx in range(3):
            logging.info(f"main_run idx: {_idx}")
            all_asin_dict = self.run_asin(zipcode, keywords, total_page_nums)
            self.quit_driver()

            if all_asin_dict:
                self.save_excel(all_asin_dict, xlsx_file)
                status = "success"
                break

        logging.info(f"======== finish: {status} ========")


if __name__ == '__main__':
    config_log("run_asin_rank.log")

    _zipcode = "77429"
    _keywords = "pack and play mattress"

    today_str = datetime.datetime.now().strftime("%Y%m%d")
    _xlsx_file = DATA_PATH + f"/{_keywords}_{_zipcode}_{today_str}.xlsx"

    _total_page_nums = 3
    obj = RunAsinRank()
    obj.main_run(_zipcode, _keywords, _total_page_nums, _xlsx_file)
